"""Export tutors from the Excel system into tutors.json for the website.

Windows quick start (PowerShell in this folder):
  py -m pip install openpyxl
  py export_tutors_json.py --xlsx "Math_and_Sciences_Hub_Full_Institution_System_ENHANCED_UNPROTECTED.xlsx" --out "tutors.json" --year 2026

Notes
- Reads sheet: TUTOR_REGISTER
- Publishes only rows with a Tutor Name.
- Uses existing "Verification Code" if present (recommended for stable codes).
- Uses existing "Verification Status" if present; otherwise derives status from "Status (Active/Inactive)".
- Optional: --writeback fills missing verification codes into the Excel file and saves a copy.
"""

import argparse
import json
import re
from datetime import date
from pathlib import Path

import openpyxl


def make_display_name(full_name: str) -> str:
    name = (full_name or "").strip()
    parts = [p for p in re.split(r"\s+", name) if p]
    if len(parts) >= 2:
        return f"{parts[0][0].upper()}. {parts[-1].title()}"
    return name


def split_subjects(text: str):
    if not text:
        return []
    parts = re.split(r"\s*[,;/]\s*", str(text).strip())
    return [p.strip() for p in parts if p.strip()]


def norm(s: str) -> str:
    return str(s or "").strip()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to the Excel workbook")
    ap.add_argument("--out", default="tutors.json", help="Output JSON file")
    ap.add_argument("--year", type=int, default=date.today().year, help="Year in verification code")
    ap.add_argument(
        "--writeback",
        action="store_true",
        help="Write missing Verification Codes back into the Excel file and save a copy",
    )
    ap.add_argument(
        "--writeback_out",
        default="",
        help="Output path for writeback copy (defaults to <xlsx>_WITH_CODES.xlsx)",
    )
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)

    if "TUTOR_REGISTER" not in wb.sheetnames:
        raise SystemExit("Sheet 'TUTOR_REGISTER' not found.")

    ws = wb["TUTOR_REGISTER"]
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    def get(row, header):
        idx = col.get(header)
        return ws.cell(row, idx).value if idx else None

    def setv(row, header, value):
        idx = col.get(header)
        if idx:
            ws.cell(row, idx).value = value

    tutors = []
    seq = 1
    missing_code_rows = []

    for r in range(2, ws.max_row + 1):
        name = get(r, "Tutor Name")
        if not name or not str(name).strip():
            continue

        subject_spec = get(r, "Subject Specialty")
        qualification = get(r, "Highest Qualification")
        institution = get(r, "University/Institution")

        active_raw = norm(get(r, "Status (Active/Inactive)")).lower()
        ver_status_raw = norm(get(r, "Verification Status"))
        ver_code_raw = norm(get(r, "Verification Code"))

        # Stable code: prefer the Excel code
        if ver_code_raw:
            code = ver_code_raw
        else:
            code = f"MSH-TUT-{args.year}-{seq:03d}"
            missing_code_rows.append((r, str(name).strip(), code))

        # Status: prefer verification status; else derive
        if ver_status_raw:
            status = "Verified" if ver_status_raw.strip().lower() == "verified" else "Pending"
        else:
            status = "Verified" if active_raw == "active" else "Pending"

        role = f"{subject_spec} Tutor" if subject_spec else "Tutor"
        subjects = split_subjects(subject_spec)

        tutors.append(
            {
                "code": code,
                "displayName": make_display_name(str(name)),
                "name": str(name).strip(),
                "role": str(role).strip(),
                "qualification": (str(qualification).strip() if qualification else ""),
                "institution": (str(institution).strip() if institution else ""),
                "subjects": subjects,
                "status": status,
            }
        )
        seq += 1

    # Write tutors.json
    out_path = Path(args.out)
    out_path.write_text(json.dumps(tutors, ensure_ascii=False, indent=2), encoding="utf-8")

    # Optional writeback: fill missing verification codes and save a copy
    if args.writeback and missing_code_rows:
        for r, _name, code in missing_code_rows:
            setv(r, "Verification Code", code)
            # If Verification Status is empty but tutor is Active, set Verified by default
            if not norm(get(r, "Verification Status")) and norm(get(r, "Status (Active/Inactive)")).lower() == "active":
                setv(r, "Verification Status", "Verified")

        wb_out = Path(args.writeback_out) if args.writeback_out else xlsx_path.with_name(xlsx_path.stem + "_WITH_CODES.xlsx")
        wb.save(wb_out)
        print(f"Writeback saved -> {wb_out}")

    # Console summary
    print(f"Exported {len(tutors)} tutors -> {out_path}")
    if missing_code_rows and not args.writeback:
        print("NOTE: Some tutors had no Verification Code in Excel. Generated codes were exported, but will change if row order changes.")
        print("      Recommended: run once with --writeback to store codes in Excel.")


if __name__ == "__main__":
    main()
